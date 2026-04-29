import streamlit as st
import anthropic
import openpyxl
from openpyxl.styles import Font
import io
import re
import copy
import os

st.set_page_config(
    page_title="GWS Quote Generator",
    page_icon="🏠",
    layout="centered"
)

# ── Styles ────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@600;700&family=Inter:wght@400;500&display=swap');

html, body, [data-testid="stAppViewContainer"] {
    background: #0f0f0f;
    color: #e8e8e8;
}
[data-testid="stAppViewContainer"] {
    background: linear-gradient(160deg, #0f0f0f 0%, #141a1a 100%);
}
h1 { font-family: 'Rajdhani', sans-serif; color: #c8a96e; font-size: 2.4rem; letter-spacing: 2px; margin-bottom: 0; }
h2, h3 { font-family: 'Rajdhani', sans-serif; color: #c8a96e; letter-spacing: 1px; }
p, label, div { font-family: 'Inter', sans-serif; }
.subtitle { color: #888; font-size: 0.9rem; letter-spacing: 3px; text-transform: uppercase; margin-top: -4px; margin-bottom: 2rem; }
.stTextArea textarea, .stTextInput input {
    background: #1c1c1c !important;
    border: 1px solid #2e2e2e !important;
    color: #e8e8e8 !important;
    border-radius: 4px !important;
    font-family: 'Inter', sans-serif !important;
}
.stTextArea textarea:focus, .stTextInput input:focus {
    border-color: #c8a96e !important;
    box-shadow: 0 0 0 1px #c8a96e30 !important;
}
.stButton > button {
    background: #c8a96e;
    color: #0f0f0f;
    border: none;
    font-family: 'Rajdhani', sans-serif;
    font-weight: 700;
    font-size: 1rem;
    letter-spacing: 2px;
    text-transform: uppercase;
    padding: 0.6rem 2rem;
    border-radius: 2px;
    transition: all 0.2s;
    width: 100%;
}
.stButton > button:hover { background: #dfc080; transform: translateY(-1px); }
.preview-box {
    background: #161c1c;
    border: 1px solid #2a2a2a;
    border-left: 3px solid #c8a96e;
    border-radius: 4px;
    padding: 1.2rem 1.4rem;
    font-family: 'Inter', monospace;
    font-size: 0.85rem;
    line-height: 1.7;
    white-space: pre-wrap;
    color: #ccc;
    margin: 1rem 0;
}
.error-box {
    background: #1a0f0f;
    border: 1px solid #5a2020;
    border-left: 3px solid #e05555;
    border-radius: 4px;
    padding: 1rem 1.4rem;
    color: #e09090;
    margin: 0.5rem 0;
}
.info-box {
    background: #0f161a;
    border: 1px solid #1e3a4a;
    border-left: 3px solid #4a9abb;
    border-radius: 4px;
    padding: 1rem 1.4rem;
    color: #90c0d0;
    margin: 0.5rem 0;
    font-size: 0.88rem;
}
.divider { border: none; border-top: 1px solid #222; margin: 1.5rem 0; }
[data-testid="stDownloadButton"] > button {
    background: #1e2e1e;
    color: #7ec87e;
    border: 1px solid #3a6a3a;
    font-family: 'Rajdhani', sans-serif;
    font-weight: 600;
    letter-spacing: 1px;
    border-radius: 2px;
    width: 100%;
}
[data-testid="stDownloadButton"] > button:hover { background: #253525; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("<h1>GWS QUOTE GENERATOR</h1>", unsafe_allow_html=True)
st.markdown('<p class="subtitle">Roofing Estimates · Excel Output</p>', unsafe_allow_html=True)

# ── System Prompt ─────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are a quote parser for GWS Roofing. Your job is to extract structured data from dictated roofing job information and return it as JSON only — no explanation, no preamble, no markdown fences.

RETURN FORMAT (strict JSON):
{
  "estimator": "string",
  "date": "DD/MM/YYYY",
  "customer_name": "string",
  "address": "string",
  "postcode": "string",
  "guarantee": "string or null",
  "notes": "string or null",
  "items": [
    {
      "type": "subheading" | "item",
      "number": null | integer,
      "description": "string (full, unsplit)",
      "cost": null | number
    }
  ]
}

RULES:
- All text → proper sentence case (first letter always capitalised).
- Date → DD/MM/YYYY format.
- Cost: only valid if £ symbol or word "pounds" present. Numbers in descriptions (e.g. "900 wide", "2 layers") are NOT costs. Use last valid monetary value per item.
- If an item has a description but no valid cost, return {"error": "missing_cost", "item": <description>}.
- Subheadings: type=subheading, number=null, cost=null.
- Items: type=item, number=integer, cost=number.
- guarantee and notes: only include if explicitly dictated, else null.
- Do not carry costs between items.
- Return ONLY valid JSON. No text before or after."""

# ── Helpers ───────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_1 = os.path.join(BASE_DIR, "GWS_Quote_-_1_Page.xlsx")
TEMPLATE_2 = os.path.join(BASE_DIR, "GWS_Quote_-_2_Page_.xlsx")


def sentence_case(text):
    if not text:
        return text
    return text[0].upper() + text[1:]


def split_description(description, max_chars=80):
    """Split a description into lines of max_chars, never breaking words."""
    words = description.split()
    lines = []
    current = ""
    for word in words:
        test = (current + " " + word).strip() if current else word
        if len(test) <= max_chars:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def parse_with_claude(dictation):
    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": dictation}]
    )
    raw = response.content[0].text.strip()
    # Strip markdown fences if present
    raw = re.sub(r"^```json\s*|^```\s*|\s*```$", "", raw, flags=re.MULTILINE).strip()
    import json
    return json.loads(raw)


def calculate_rows_needed(items):
    """Calculate how many rows the item layout will occupy (rows 9+)."""
    row = 9
    prev_was_item = False
    for idx, item in enumerate(items):
        if item["type"] == "subheading":
            if prev_was_item:
                row += 1  # blank row after item before subheading
            row += 1  # subheading row
            prev_was_item = False
        else:
            lines = split_description(item["description"])
            row += len(lines)
            row += 1  # blank row after item
            prev_was_item = True
    return row - 1  # last used row


def write_quote_to_excel(data):
    """Load correct template, write all data, return bytes."""
    # Determine template
    rows_needed = calculate_rows_needed(data["items"])
    use_2page = rows_needed > 34

    template_path = TEMPLATE_2 if use_2page else TEMPLATE_1
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    max_item_row = 85 if use_2page else 34
    guarantee_row = 91 if use_2page else 40
    notes_row = 92 if use_2page else 41

    # Header fields
    ws["A1"] = f"Estimator :  {sentence_case(data['estimator'])}"
    ws["C3"] = data["date"]
    ws["C4"] = sentence_case(data["customer_name"])
    ws["C5"] = sentence_case(data["address"])
    ws["C6"] = data["postcode"].upper()

    # Items
    current_row = 9
    prev_was_item = False

    for item in data["items"]:
        if item["type"] == "subheading":
            if prev_was_item:
                current_row += 1  # blank row before subheading
            cell = ws.cell(row=current_row, column=2)
            cell.value = sentence_case(item["description"])
            cell.font = Font(bold=True, name=cell.font.name or "Arial")
            current_row += 1
            prev_was_item = False

        else:  # item
            lines = split_description(item["description"])
            item_num = item["number"]
            cost = item["cost"]

            for i, line in enumerate(lines):
                is_first = i == 0
                is_last = i == len(lines) - 1

                if is_first:
                    ws.cell(row=current_row, column=1).value = item_num
                # Column B
                ws.cell(row=current_row, column=2).value = sentence_case(line) if is_first else line
                # Cost on last row
                if is_last:
                    ws.cell(row=current_row, column=7).value = cost
                current_row += 1

            current_row += 1  # blank row after item
            prev_was_item = True

    # Guarantee / Notes
    if data.get("guarantee"):
        ws.cell(row=guarantee_row, column=3).value = sentence_case(data["guarantee"])
    if data.get("notes"):
        ws.cell(row=notes_row, column=3).value = sentence_case(data["notes"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), use_2page


def build_preview(data):
    lines = []
    lines.append(f"Estimator : {data['estimator']}")
    lines.append(f"Date      : {data['date']}")
    lines.append(f"Customer  : {data['customer_name']}")
    lines.append(f"Address   : {data['address']}")
    lines.append(f"Postcode  : {data['postcode']}")
    if data.get("guarantee"):
        lines.append(f"Guarantee : {data['guarantee']}")
    if data.get("notes"):
        lines.append(f"Notes     : {data['notes']}")
    lines.append("")
    lines.append("─" * 50)
    for item in data["items"]:
        if item["type"] == "subheading":
            lines.append(f"\n  ▶ {item['description'].upper()}")
        else:
            cost_str = f"£{item['cost']:,.2f}" if item["cost"] else "NO COST"
            lines.append(f"  Item {item['number']} – {item['description']} – {cost_str}")
    return "\n".join(lines)


# ── State ──────────────────────────────────────────────────────────────────────
if "parsed_data" not in st.session_state:
    st.session_state.parsed_data = None
if "preview_text" not in st.session_state:
    st.session_state.preview_text = None
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "filename" not in st.session_state:
    st.session_state.filename = None
if "error" not in st.session_state:
    st.session_state.error = None

# ── Input ──────────────────────────────────────────────────────────────────────
st.markdown("### Dictation Input")
st.markdown(
    '<div class="info-box">Begin with <strong>New quote</strong> to start fresh. '
    'Dictate estimator, date, customer, address, postcode, then items. '
    'Costs must include £ or "pounds".</div>',
    unsafe_allow_html=True
)

dictation = st.text_area(
    "Enter or paste dictation:",
    height=220,
    placeholder=(
        "New quote\n"
        "Estimator John Smith\n"
        "Date 17th February 2026\n"
        "Name of customer Mr A Jones\n"
        "Address 14 High Street, Guildford\n"
        "Postcode GU1 3AA\n\n"
        "Item 1 Strip and re-tile front roof slope in concrete interlocking tiles £2500\n"
        "Next item Replace all lead flashings to chimney stack £850"
    ),
    label_visibility="collapsed"
)

col1, col2 = st.columns([3, 1])
with col1:
    parse_btn = st.button("🔍 Parse & Preview", use_container_width=True)
with col2:
    if st.button("↺ Reset", use_container_width=True):
        st.session_state.parsed_data = None
        st.session_state.preview_text = None
        st.session_state.excel_bytes = None
        st.session_state.filename = None
        st.session_state.error = None
        st.rerun()

# ── Parse ──────────────────────────────────────────────────────────────────────
if parse_btn and dictation.strip():
    with st.spinner("Parsing dictation…"):
        try:
            data = parse_with_claude(dictation)

            if "error" in data:
                st.session_state.error = f"Missing cost for item: {data.get('item', 'unknown')}"
                st.session_state.parsed_data = None
            else:
                # Validate required fields
                missing = []
                for field in ["estimator", "date", "customer_name", "address", "postcode"]:
                    if not data.get(field):
                        missing.append(field)
                if missing:
                    st.session_state.error = f"Missing required fields: {', '.join(missing)}"
                else:
                    st.session_state.parsed_data = data
                    st.session_state.preview_text = build_preview(data)
                    st.session_state.error = None
                    st.session_state.excel_bytes = None

        except Exception as e:
            st.session_state.error = f"Parsing error: {str(e)}"

# ── Error ─────────────────────────────────────────────────────────────────────
if st.session_state.error:
    st.markdown(
        f'<div class="error-box">⚠ {st.session_state.error}</div>',
        unsafe_allow_html=True
    )

# ── Preview ───────────────────────────────────────────────────────────────────
if st.session_state.preview_text:
    st.markdown("<hr class='divider'>", unsafe_allow_html=True)
    st.markdown("### Quote Preview")
    st.markdown(
        f'<div class="preview-box">{st.session_state.preview_text}</div>',
        unsafe_allow_html=True
    )

    confirm_col, _ = st.columns([2, 1])
    with confirm_col:
        if st.button("✅ Confirm & Generate Excel", use_container_width=True):
            with st.spinner("Generating Excel file…"):
                try:
                    excel_bytes, used_2page = write_quote_to_excel(st.session_state.parsed_data)
                    address = st.session_state.parsed_data.get("address", "Quote")
                    # Sanitise address for filename
                    safe_addr = re.sub(r"[^\w\s-]", "", address).strip()
                    safe_addr = re.sub(r"\s+", " ", safe_addr)
                    st.session_state.filename = f"GWS Quote {safe_addr}.xlsx"
                    st.session_state.excel_bytes = excel_bytes
                    template_label = "2-page" if used_2page else "1-page"
                    st.success(f"Excel generated using {template_label} template.")
                except Exception as e:
                    st.session_state.error = f"Excel generation error: {str(e)}"
                    st.rerun()

# ── Download ──────────────────────────────────────────────────────────────────
if st.session_state.excel_bytes:
    st.markdown("<hr class='divider'>", unsafe_allow_html=True)
    st.download_button(
        label=f"⬇ Download  {st.session_state.filename}",
        data=st.session_state.excel_bytes,
        file_name=st.session_state.filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
